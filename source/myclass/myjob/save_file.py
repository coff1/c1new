from ..mydb import mydb

import openpyxl

def save_query_results_to_excel(field_names_list, field_values_list, file_name="query_results.xlsx",pagename=[]):
    """
    Saves multiple database query results to an Excel file, with an unlimited number of rows per sheet, and expands the table with a horizontal line when the characters in a cell are too long to fit in the column.

    Parameters:
        - field_names_list: a list of field names for each query result
        - field_values_list: a list of field values for each query result
        - file_name: the name of the Excel file to save the results to (default: "query_results.xlsx")

    Returns: None
    """

    # Create a new Excel workbook
    wb = openpyxl.Workbook()

    # Loop through the query results
    for i, (field_names, field_values) in enumerate(zip(field_names_list, field_values_list)):
        # print(i)

        # Create a new worksheet for the query result
        try:
            name = "Result Of {}".format(pagename[i])
            
        except BaseException:
            name = "Query Result {}".format(i)

        ws = wb.create_sheet(name)

        # Set the column width based on the maximum length of the data in each column
        for j, field_name in enumerate(field_names):
            max_length = 20
            column_letter = openpyxl.utils.get_column_letter(j+1)
            ws.column_dimensions[column_letter].width = max_length

        # Write the field names to the first row
        for j, field_name in enumerate(field_names):
            cell = ws.cell(row=1, column=j+1)
            cell.value = field_name
            # 会自动换行
            # cell.alignment = openpyxl.styles.Alignment(wrap_text=True, shrink_to_fit=True)
            cell.alignment = openpyxl.styles.Alignment(shrink_to_fit=True)

        # Write the field values to the remaining rows
        for k, field_value_row in enumerate(field_values):
            for l, field_value in enumerate(field_value_row):
                cell = ws.cell(row=k+2, column=l+1)
                cell.value = field_value
                # cell.alignment = openpyxl.styles.Alignment(wrap_text=True, shrink_to_fit=True)
                cell.alignment = openpyxl.styles.Alignment(shrink_to_fit=True)

    # Remove the default sheet that is created when the workbook is first created
    wb.remove(wb["Sheet"])

    # Save the Excel file
    wb.save(file_name)